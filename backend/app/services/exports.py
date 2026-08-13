"""Payslip PDFs and the monthly Excel export.

The payslip deliberately prints the day-by-day derivation underneath the money.
When someone disputes their overtime, the answer has to be on the paper they
are holding, not in a screen only the owner can see.
"""

from __future__ import annotations

import calendar
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..config import settings

# ReportLab's built-in fonts have no rupee glyph, so the symbol would print as a
# black box. "Rs." is used on PDFs; the web UI shows the real symbol.
PDF_CURRENCY = "Rs."

BRAND = colors.HexColor("#1f3a5f")
MUTED = colors.HexColor("#6b7280")
LINE = colors.HexColor("#d1d5db")
BAND = colors.HexColor("#f3f4f6")

STATUS_LABELS = {
    "FULL": "Present",
    "HALF": "Half day",
    "LEAVE": "Leave",
    "PAID_LEAVE": "Paid leave",
    "WEEKOFF": "Week off",
    "HOLIDAY": "Holiday",
}


def money(value) -> str:
    return f"{PDF_CURRENCY} {Decimal(str(value)):,.2f}"


def period_label(year: int, month: int) -> str:
    return f"{calendar.month_name[month]} {year}"


# --- payslip PDF ------------------------------------------------------------


def _styles():
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "title", parent=base["Title"], fontSize=16, textColor=BRAND, spaceAfter=2
        ),
        "sub": ParagraphStyle(
            "sub", parent=base["Normal"], fontSize=9.5, textColor=MUTED
        ),
        "h2": ParagraphStyle(
            "h2",
            parent=base["Heading2"],
            fontSize=10.5,
            textColor=BRAND,
            spaceBefore=10,
            spaceAfter=4,
        ),
        "small": ParagraphStyle(
            "small", parent=base["Normal"], fontSize=8, textColor=MUTED
        ),
        "note": ParagraphStyle(
            "note", parent=base["Normal"], fontSize=8.5, textColor=colors.HexColor("#9a3412")
        ),
    }


def _payslip_flowables(row: dict, year: int, month: int, styles: dict) -> list:
    flow: list = []

    flow.append(Paragraph(settings.business_name, styles["title"]))
    flow.append(Paragraph(f"Salary slip for {period_label(year, month)}", styles["sub"]))
    flow.append(Spacer(1, 8))

    details = Table(
        [
            ["Employee", row["employee_name"], "Code", row["employee_code"]],
            ["Location", row["location"], "Monthly salary", money(row["monthly_salary"])],
            [
                "Daily rate",
                money(row["daily_rate"]),
                "Days present",
                str(row["days_full"]),
            ],
        ],
        colWidths=[28 * mm, 55 * mm, 32 * mm, 45 * mm],
    )
    details.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TEXTCOLOR", (0, 0), (0, -1), MUTED),
                ("TEXTCOLOR", (2, 0), (2, -1), MUTED),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
                ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("LINEBELOW", (0, 0), (-1, -2), 0.4, LINE),
            ]
        )
    )
    flow.append(details)

    # --- attendance summary ---
    flow.append(Paragraph("Attendance", styles["h2"]))
    summary = Table(
        [
            ["Present", "Half days", "Leaves", "Paid leave", "Week offs", "Nights"],
            [
                str(row["days_full"]),
                str(row["days_half"]),
                str(row["days_leave"]),
                str(row["days_paid_leave"]),
                str(row["days_weekoff"]),
                str(row["nights"]),
            ],
        ],
        colWidths=[26.6 * mm] * 6,
    )
    summary.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), BAND),
                ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, LINE),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    flow.append(summary)

    ot_note = (
        f"Overtime worked {row['ot_hours_raw']}, paid on {row['ot_hours_paid']}."
    )
    if row.get("ot_adjustment") and row["ot_adjustment"] != "0:00":
        ot_note += f" Night adjustment of {row['ot_adjustment']} applied."
    if row["nights"]:
        ot_note += " Overtime on night days is paid as night duty, not as overtime."
    flow.append(Spacer(1, 4))
    flow.append(Paragraph(ot_note, styles["small"]))

    # --- money ---
    flow.append(Paragraph("Salary calculation", styles["h2"]))
    # The advance is deliberately NOT listed here: it has its own line in the
    # totals block below, and showing it in both places reads as a double
    # deduction to the person holding the slip.
    lines = [
        ["Earnings", "", "Deductions", ""],
        ["Basic salary", money(row["monthly_salary"]), "Leave deduction", money(row["leave_deduction"])],
        ["Overtime", money(row["ot_pay"]), "Half-day deduction", money(row["halfday_deduction"])],
        ["Night duty", money(row["night_pay"]), "", ""],
        ["Attendance bonus", money(row["attendance_bonus"]), "", ""],
    ]
    if Decimal(row["adjustments_total"]) != 0:
        lines.append(["Adjustments", money(row["adjustments_total"]), "", ""])

    table = Table(lines, colWidths=[40 * mm, 40 * mm, 40 * mm, 40 * mm])
    table.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), BAND),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, 0), (-1, 0), BRAND),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, LINE),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    flow.append(table)

    totals = Table(
        [
            ["Grand total", money(row["grand_total"])],
            ["Less advance", money(row["advances_deducted"])],
            ["Net payable", money(row["net_payable"])],
        ],
        colWidths=[120 * mm, 40 * mm],
    )
    totals.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("LINEABOVE", (0, 0), (-1, 0), 0.4, LINE),
                ("BACKGROUND", (0, 2), (-1, 2), BRAND),
                ("TEXTCOLOR", (0, 2), (-1, 2), colors.white),
                ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
                ("FONTSIZE", (0, 2), (-1, 2), 11.5),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    flow.append(Spacer(1, 6))
    flow.append(totals)

    if row.get("bonus_reason"):
        granted = "granted" if row.get("bonus_granted") else "not granted"
        flow.append(Spacer(1, 4))
        flow.append(
            Paragraph(f"Attendance bonus {granted}: {row['bonus_reason']}", styles["small"])
        )

    for note in row.get("notes", []):
        flow.append(Paragraph(note, styles["small"]))
    for flag in row.get("flags", []):
        flow.append(Paragraph(f"Please check: {flag}", styles["note"]))

    # --- day by day ---
    breakdown = row.get("breakdown") or []
    if breakdown:
        flow.append(Paragraph("Day by day", styles["h2"]))
        header = ["Date", "Status", "Worked", "OT", "Night"]
        body = [header]
        has_discarded = False
        for entry in breakdown:
            # Rule 3 throws away a day's OT if it is under 30 minutes. Printing
            # the raw figure with nothing to mark it would make the column fail
            # to add up to the overtime actually paid -- exactly the sort of
            # thing that starts an argument. Discarded time is bracketed.
            discarded = (
                not entry["is_night"]
                and entry["ot_minutes"] > 0
                and entry["counted_ot_minutes"] == 0
            )
            if discarded:
                has_discarded = True
                ot_text = "(" + entry["ot"] + ")"
            elif entry["ot"] != "0:00":
                ot_text = entry["ot"]
            else:
                ot_text = "-"

            body.append(
                [
                    entry["date"][-2:],
                    STATUS_LABELS.get(entry["status"], entry["status"]),
                    entry["worked"],
                    ot_text,
                    "Yes" if entry["is_night"] else "-",
                ]
            )

        # Two columns side by side so a 31-day month fits on one page.
        half = (len(body) + 1) // 2
        left, right = body[:half], [header] + body[half:]
        left_rows, right_rows = len(left), len(right)
        while len(right) < len(left):
            right.append([""] * 5)
        while len(left) < len(right):
            left.append([""] * 5)

        merged = [l + [""] + r for l, r in zip(left, right)]
        widths = [10 * mm, 20 * mm, 14 * mm, 12 * mm, 12 * mm, 6 * mm] + [
            10 * mm,
            20 * mm,
            14 * mm,
            12 * mm,
            12 * mm,
        ]
        grid = Table(merged, colWidths=widths, repeatRows=1)
        grid.setStyle(
            TableStyle(
                [
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("BACKGROUND", (0, 0), (4, 0), BAND),
                    ("BACKGROUND", (6, 0), (10, 0), BAND),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    # Grid only the rows that hold real days, so the shorter
                    # column does not trail an empty box.
                    ("GRID", (0, 0), (4, left_rows - 1), 0.3, LINE),
                    ("GRID", (6, 0), (10, right_rows - 1), 0.3, LINE),
                    ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ]
            )
        )
        flow.append(grid)

        if has_discarded:
            flow.append(Spacer(1, 3))
            flow.append(
                Paragraph(
                    "Overtime shown in brackets was under 30 minutes for that "
                    "day and is not counted, as per the overtime rules.",
                    styles["small"],
                )
            )

    flow.append(Spacer(1, 14))
    flow.append(
        Paragraph(
            "Employee signature: _______________________ &nbsp;&nbsp;&nbsp; "
            "Manager signature: _______________________",
            styles["small"],
        )
    )
    return flow


def payslip_pdf(rows: list[dict], year: int, month: int) -> bytes:
    """One payslip per employee, one page each."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Payslips {period_label(year, month)}",
    )
    styles = _styles()

    flow: list = []
    for index, row in enumerate(rows):
        if index:
            flow.append(PageBreak())
        flow.extend(_payslip_flowables(row, year, month, styles))

    doc.build(flow)
    return buffer.getvalue()


# --- Excel export -----------------------------------------------------------

EXPORT_COLUMNS = [
    ("Code", "employee_code"),
    ("Name", "employee_name"),
    ("Location", "location"),
    ("Salary", "monthly_salary"),
    ("Present", "days_full"),
    ("Half days", "days_half"),
    ("Leaves", "days_leave"),
    ("Paid leave", "days_paid_leave"),
    ("Week offs", "days_weekoff"),
    ("Nights", "nights"),
    ("OT worked", "ot_hours_raw"),
    ("OT paid on", "ot_hours_paid"),
    ("OT pay", "ot_pay"),
    ("Night pay", "night_pay"),
    ("Bonus", "attendance_bonus"),
    ("Leave deduction", "leave_deduction"),
    ("Half-day deduction", "halfday_deduction"),
    ("Adjustments", "adjustments_total"),
    ("Grand total", "grand_total"),
    ("Advance", "advances_deducted"),
    ("Net payable", "net_payable"),
]

NUMERIC_KEYS = {
    "monthly_salary",
    "ot_pay",
    "night_pay",
    "attendance_bonus",
    "leave_deduction",
    "halfday_deduction",
    "adjustments_total",
    "grand_total",
    "advances_deducted",
    "net_payable",
}


def payroll_xlsx(payload: dict) -> bytes:
    """The whole month as a spreadsheet, plus a day-by-day sheet."""
    year, month = payload["year"], payload["month"]
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Salary"
    sheet.append([f"{settings.business_name} - Salary for {period_label(year, month)}"])
    sheet["A1"].font = Font(bold=True, size=13)
    sheet.append([])

    header_row = 3
    sheet.append([label for label, _ in EXPORT_COLUMNS])
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3A5F")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in payload["rows"]:
        values = []
        for _, key in EXPORT_COLUMNS:
            value = row.get(key, "")
            values.append(float(Decimal(value)) if key in NUMERIC_KEYS else value)
        sheet.append(values)

    for index, (label, key) in enumerate(EXPORT_COLUMNS, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = max(11, min(22, len(label) + 4))
        if key in NUMERIC_KEYS:
            for cell in sheet[letter][header_row:]:
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = sheet[f"A{header_row + 1}"]

    total_row = header_row + len(payload["rows"]) + 1
    sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for index, (_, key) in enumerate(EXPORT_COLUMNS, start=1):
        if key in NUMERIC_KEYS:
            letter = get_column_letter(index)
            cell = sheet.cell(row=total_row, column=index)
            cell.value = f"=SUM({letter}{header_row + 1}:{letter}{total_row - 1})"
            cell.font = Font(bold=True)
            cell.number_format = "#,##0.00"

    # --- day-by-day sheet ---
    detail = workbook.create_sheet("Day by day")
    detail.append(["Code", "Name", "Date", "Status", "Worked", "Overtime", "Night"])
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3A5F")

    for row in payload["rows"]:
        for entry in row.get("breakdown", []):
            detail.append(
                [
                    row["employee_code"],
                    row["employee_name"],
                    entry["date"],
                    STATUS_LABELS.get(entry["status"], entry["status"]),
                    entry["worked"],
                    entry["ot"],
                    "Yes" if entry["is_night"] else "",
                ]
            )
    for index, width in enumerate([10, 22, 12, 12, 10, 10, 8], start=1):
        detail.column_dimensions[get_column_letter(index)].width = width
    detail.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
